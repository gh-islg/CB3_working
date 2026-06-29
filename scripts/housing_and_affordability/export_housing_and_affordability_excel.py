"""Export the Housing & Affordability tract metrics to a documented Excel workbook."""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

PROJECT_DIR = Path(__file__).resolve().parents[2]
CLEAN_DIR = PROJECT_DIR / "data" / "clean"
INPUT_PATH = CLEAN_DIR / "housing_and_affordability.csv"
EXCEL_OUTPUT_PATH = CLEAN_DIR / "housing_and_affordability.xlsx"

clean = pd.read_csv(INPUT_PATH, dtype={"GEOID": "string"})

income_blocks = {
    "lt_10k": 2,
    "10k_19k": 11,
    "20k_34k": 20,
    "35k_49k": 29,
    "50k_74k": 38,
    "75k_99k": 47,
    "100k_plus": 56,
}
# Build a complete data dictionary directly from the calculations above. Each
# output variable receives a definition, unit, geography, source, calculation,
# and interpretation or limitation note.
dictionary_rows = []


def add_dictionary_row(
    variable_name,
    category,
    description,
    unit,
    geography,
    source_file,
    source_fields,
    calculation,
    notes="",
):
    dictionary_rows.append(
        {
            "variable_name": variable_name,
            "category": category,
            "description": description,
            "unit": unit,
            "geography": geography,
            "source_file": source_file,
            "source_fields": source_fields,
            "calculation": calculation,
            "notes_and_limitations": notes,
        }
    )


# Geography and identifiers.
geography_dictionary = {
    "GEOID": (
        "Unique 11-character 2020 Census tract identifier.",
        "Identifier",
        "2020 Census tract",
        "nyc_2020_census_tract_nta_cdta_relationships.xlsx",
        "GEOID",
        "Direct field; stored as text to preserve leading zeros.",
    ),
    "tract_label": (
        "Human-readable Census tract number.",
        "Text",
        "2020 Census tract",
        "nyc_2020_census_tract_nta_cdta_relationships.xlsx",
        "CTLabel",
        "Direct field.",
    ),
    "tract_name": (
        "Full Census tract name.",
        "Text",
        "2020 Census tract",
        "nyc_2020_census_tract_nta_cdta_relationships.xlsx",
        "CTLabel, BoroName",
        "Constructed as 'Census Tract {CTLabel}, {BoroName} County, New York'.",
    ),
    "nta_code": (
        "2020 Neighborhood Tabulation Area code.",
        "Identifier",
        "2020 NTA",
        "nyc_2020_census_tract_nta_cdta_relationships.xlsx",
        "NTACode",
        "Direct field associated with the tract.",
    ),
    "nta_name": (
        "2020 Neighborhood Tabulation Area name.",
        "Text",
        "2020 NTA",
        "nyc_2020_census_tract_nta_cdta_relationships.xlsx",
        "NTAName",
        "Direct field associated with the tract.",
    ),
    "cdta_code": (
        "2020 Community District Tabulation Area code.",
        "Identifier",
        "2020 CDTA",
        "nyc_2020_census_tract_nta_cdta_relationships.xlsx",
        "CDTACode",
        "Direct field; all output rows are MN03.",
    ),
    "cdta_name": (
        "2020 Community District Tabulation Area name.",
        "Text",
        "2020 CDTA",
        "nyc_2020_census_tract_nta_cdta_relationships.xlsx",
        "CDTAName",
        "Direct field associated with the tract.",
    ),
}
for variable_name, values in geography_dictionary.items():
    add_dictionary_row(variable_name, "Geography", *values)


# ACS severe rent burden.
add_dictionary_row(
    "severe_rent_burden_households",
    "Rent burden",
    "Renter households spending 50% or more of household income on gross rent.",
    "Households",
    "2020 Census tract",
    "acs_5yr_2024_B25070.csv",
    "B25070_010E",
    "Direct ACS estimate.",
)
add_dictionary_row(
    "rent_burden_denominator",
    "Rent burden",
    "Renter households for whom gross rent as a percentage of income was computed.",
    "Households",
    "2020 Census tract",
    "acs_5yr_2024_B25070.csv",
    "B25070_001E, B25070_011E",
    "B25070_001E minus B25070_011E (not computed).",
)
add_dictionary_row(
    "severe_rent_burden_pct",
    "Rent burden",
    "Share of renter households with computed rent burden spending 50% or more of income on gross rent.",
    "Percent",
    "2020 Census tract",
    "acs_5yr_2024_B25070.csv",
    "B25070_010E, B25070_001E, B25070_011E",
    "100 × B25070_010E / (B25070_001E - B25070_011E).",
    "Blank where the denominator is zero.",
)


# ACS crowding.
add_dictionary_row(
    "crowded_households",
    "Crowding",
    "Occupied households with more than one occupant per room.",
    "Households",
    "2020 Census tract",
    "acs_5yr_2024_B25014.csv",
    "B25014_005E–007E and B25014_011E–013E",
    "Sum of owner and renter cells representing 1.01–1.50, 1.51–2.00, and 2.01 or more occupants per room.",
)
add_dictionary_row(
    "occupied_households",
    "Crowding",
    "Total occupied housing units used as the crowding denominator.",
    "Households",
    "2020 Census tract",
    "acs_5yr_2024_B25014.csv",
    "B25014_001E",
    "Direct ACS estimate.",
)
add_dictionary_row(
    "crowded_households_pct",
    "Crowding",
    "Share of occupied households with more than one occupant per room.",
    "Percent",
    "2020 Census tract",
    "acs_5yr_2024_B25014.csv",
    "B25014_001E, B25014_005E–007E, B25014_011E–013E",
    "100 × crowded_households / occupied_households.",
    "Blank where the denominator is zero.",
)


# ACS tenure overall.
tenure_base_dictionary = {
    "occupied_housing_units": (
        "Total occupied housing units.",
        "Housing units",
        "B25003_001E",
        "Direct ACS estimate.",
    ),
    "owner_occupied_units": (
        "Owner-occupied housing units.",
        "Housing units",
        "B25003_002E",
        "Direct ACS estimate.",
    ),
    "renter_occupied_units": (
        "Renter-occupied housing units.",
        "Housing units",
        "B25003_003E",
        "Direct ACS estimate.",
    ),
    "owner_occupied_pct": (
        "Share of occupied housing units that are owner occupied.",
        "Percent",
        "B25003_001E, B25003_002E",
        "100 × B25003_002E / B25003_001E.",
    ),
    "renter_occupied_pct": (
        "Share of occupied housing units that are renter occupied.",
        "Percent",
        "B25003_001E, B25003_003E",
        "100 × B25003_003E / B25003_001E.",
    ),
}
for variable_name, values in tenure_base_dictionary.items():
    add_dictionary_row(
        variable_name,
        "Tenure",
        values[0],
        values[1],
        "2020 Census tract",
        "acs_5yr_2024_B25003.csv",
        values[2],
        values[3],
        "Percentage is blank where the group denominator is zero."
        if variable_name.endswith("_pct")
        else "",
    )


# ACS tenure by race/ethnicity.
race_dictionary = {
    "white_alone": ("A", "White alone"),
    "black_alone": ("B", "Black or African American alone"),
    "asian_alone": ("D", "Asian alone"),
    "white_non_hispanic": ("H", "White alone, not Hispanic or Latino"),
    "hispanic": ("I", "Hispanic or Latino"),
}
for label, (suffix, group_name) in race_dictionary.items():
    for tenure_type, cell in [("owner", "002"), ("renter", "003")]:
        variable_name = f"{tenure_type}_occupied_pct_{label}"
        add_dictionary_row(
            variable_name,
            "Tenure by race/ethnicity",
            f"Share of occupied housing units for {group_name} householders that are {tenure_type} occupied.",
            "Percent",
            "2020 Census tract",
            "acs_5yr_2024_B25003.csv",
            f"B25003{suffix}_001E, B25003{suffix}_{cell}E",
            f"100 × B25003{suffix}_{cell}E / B25003{suffix}_001E.",
            "Householder race/ethnicity classification; blank where the group denominator is zero.",
        )


# ACS rent burden by income band.
income_labels = {
    "lt_10k": "less than $10,000",
    "10k_19k": "$10,000–$19,999",
    "20k_34k": "$20,000–$34,999",
    "35k_49k": "$35,000–$49,999",
    "50k_74k": "$50,000–$74,999",
    "75k_99k": "$75,000–$99,999",
    "100k_plus": "$100,000 or more",
}
for label, start_cell in income_blocks.items():
    total_field = f"B25074_{start_cell:03d}E"
    burden_fields = ", ".join(
        f"B25074_{cell:03d}E" for cell in range(start_cell + 4, start_cell + 8)
    )
    not_computed_field = f"B25074_{start_cell + 8:03d}E"
    band_name = income_labels[label]
    add_dictionary_row(
        f"rent_burden_30plus_households_income_{label}",
        "Rent burden by income",
        f"Renter households with income {band_name} spending at least 30% of income on gross rent.",
        "Households",
        "2020 Census tract",
        "acs_5yr_2024_B25074.csv",
        burden_fields,
        "Sum of the 30–34.9%, 35–39.9%, 40–49.9%, and 50% or more cells.",
    )
    add_dictionary_row(
        f"rent_burden_denominator_income_{label}",
        "Rent burden by income",
        f"Renter households with income {band_name} for whom rent burden was computed.",
        "Households",
        "2020 Census tract",
        "acs_5yr_2024_B25074.csv",
        f"{total_field}, {not_computed_field}",
        f"{total_field} minus {not_computed_field} (not computed).",
    )
    add_dictionary_row(
        f"rent_burden_30plus_pct_income_{label}",
        "Rent burden by income",
        f"Share of renter households with income {band_name} spending at least 30% of income on gross rent.",
        "Percent",
        "2020 Census tract",
        "acs_5yr_2024_B25074.csv",
        f"{total_field}, {burden_fields}, {not_computed_field}",
        "100 × 30%-or-more household count / computed-rent-burden denominator.",
        "Blank where the denominator is zero.",
    )


# HPD housing maintenance code violations.
hpd_classes = {
    "hpd_open_violations": ("all classes", "All open violations"),
    "hpd_open_class_a": ("Class A", "Open non-hazardous Class A violations"),
    "hpd_open_class_b": ("Class B", "Open hazardous Class B violations"),
    "hpd_open_class_c": ("Class C", "Open immediately hazardous Class C violations"),
    "hpd_open_class_bc": (
        "Classes B and C",
        "Open hazardous or immediately hazardous Class B/C violations",
    ),
}
for variable_name, (class_filter, description) in hpd_classes.items():
    add_dictionary_row(
        variable_name,
        "Housing quality",
        description,
        "Violations",
        "2020 Census tract",
        "Housing_Maintenance_Code_Violations_20260504.csv",
        "ViolationStatus, Class, Latitude, Longitude",
        f"Filter valid Manhattan CB3 records to ViolationStatus='Open'; count {class_filter}; spatially join coordinates to 2020 tract.",
        "Counts violations, not unique buildings or units.",
    )


# Executed evictions.
for variable_name, year_label in [
    ("executed_evictions_total", "all dates in the local file"),
    ("executed_evictions_2024", "calendar year 2024"),
    ("executed_evictions_2025", "calendar year 2025"),
]:
    add_dictionary_row(
        variable_name,
        "Evictions",
        f"Residential marshal evictions executed during {year_label}.",
        "Executed evictions",
        "2020 Census tract",
        "Evictions_20260420.csv",
        "Residential/Commercial, Executed Date, Latitude, Longitude",
        f"Filter Borough='MANHATTAN', Community Board=3, Residential/Commercial='Residential'; filter to {year_label}; spatially join coordinates to 2020 tract.",
        "Represents executed evictions, not eviction filings.",
    )


# Furman subsidized housing and recent construction.
subsidized_dictionary = {
    "subsidized_properties": (
        "Subsidized housing tax lots/properties.",
        "Properties",
        "bbl, latitude, longitude",
        "Count one BBL record after filtering cd_id=103 and spatially joining its coordinates.",
    ),
    "subsidized_units": (
        "Residential units on subsidized housing properties.",
        "Housing units",
        "res_units",
        "Sum res_units by spatially assigned tract.",
    ),
    "subsidized_units_expiring_2025_2030": (
        "Units on properties whose earliest recorded subsidy end year is 2025–2030.",
        "Housing units",
        "res_units and all end_* fields",
        "Find minimum valid year across property end_* fields; sum res_units when earliest year is 2025–2030.",
    ),
    "subsidized_units_expiring_2031_2040": (
        "Units on properties whose earliest recorded subsidy end year is 2031–2040.",
        "Housing units",
        "res_units and all end_* fields",
        "Find minimum valid year across property end_* fields; sum res_units when earliest year is 2031–2040.",
    ),
    "senior_subsidized_properties": (
        "Subsidized properties flagged for Section 202/8 or PRAC 202 senior housing.",
        "Properties",
        "prog_202_8, prog_prac_202",
        "Count property when either program flag equals 1.",
    ),
    "senior_subsidized_units": (
        "Units on subsidized properties flagged for Section 202/8 or PRAC 202 senior housing.",
        "Housing units",
        "res_units, prog_202_8, prog_prac_202",
        "Sum res_units where either senior program flag equals 1.",
    ),
}
for variable_name, values in subsidized_dictionary.items():
    add_dictionary_row(
        variable_name,
        "Subsidized housing",
        values[0],
        values[1],
        "2020 Census tract",
        "FC_SHD_bbl_analysis_2025-05-13.csv",
        values[2],
        values[3],
        "Property-level expiration estimates may not reconcile to Furman's separately published community-district totals."
        if "expiring" in variable_name
        else "",
    )

add_dictionary_row(
    "new_affordable_properties_since_2018",
    "Affordable construction",
    "Affordable properties with a new-construction subsidy start in 2018 or later.",
    "Properties",
    "2020 Census tract",
    "FC_SHD_subsidy_analysis_2025-05-13.csv",
    "ref_bbl, preservation, start_date",
    "Filter cd_id=103, preservation='New Construction', start year >=2018; deduplicate by BBL; use property tract assignment.",
)
add_dictionary_row(
    "new_affordable_units_since_2018",
    "Affordable construction",
    "Reported units at affordable properties with a new-construction subsidy start in 2018 or later.",
    "Housing units",
    "2020 Census tract",
    "FC_SHD_subsidy_analysis_2025-05-13.csv",
    "ref_bbl, preservation, start_date, tot_units",
    "Filter recent new construction; take maximum tot_units per BBL to avoid subsidy-row duplication; sum by tract.",
    "The local Furman file does not provide a defensible AMI-band variable.",
)


# Indoor environmental and NYCHA violations.
complaint_dictionary = {
    "indoor_environmental_complaints": ("all retained complaint types", "Indoor environmental complaints"),
    "indoor_air_quality_complaints": ("Indoor Air Quality", "Indoor air-quality complaints"),
    "mold_complaints": ("Mold", "Mold complaints"),
    "asbestos_complaints": ("Asbestos", "Asbestos complaints"),
    "indoor_sewage_complaints": ("Indoor Sewage", "Indoor-sewage complaints"),
}
for variable_name, (complaint_type, description) in complaint_dictionary.items():
    add_dictionary_row(
        variable_name,
        "Housing-related health",
        description,
        "Complaints",
        "2020 Census tract",
        "DOHMH_Indoor_Environmental_Complaints_20260428.csv",
        "Complaint_Type_311, Deleted, Latitude, Longitude",
        f"Filter Manhattan Community Board 3 and Deleted!='Yes'; count {complaint_type}; spatially join coordinates to 2020 tract.",
        "Complaint counts may reflect reporting behavior as well as underlying conditions.",
    )
add_dictionary_row(
    "nycha_code_violations",
    "Housing-related health",
    "NYCHA housing maintenance code violation records.",
    "Violations",
    "2020 Census tract",
    "Housing_Maintenance_Code_Violations_NYCHA_properties_20260420.csv",
    "Primary Borough Name, Community Board, Latitude, Longitude",
    "Filter Primary Borough Name='MANHATTAN' and Community Board=103; spatially join coordinates and count records.",
    "Counts violations, not unique NYCHA buildings or units.",
)


# Community-district context fields.
context_dictionary = {
    "private_eviction_filings_2024_cd_context": (
        "Private eviction filings in MN03 during 2024.",
        "Filings",
        "communitydistrict-privateevictionfilings.csv",
        "2024",
    ),
    "subsidized_units_expiring_2025_2030_cd_context": (
        "Published MN03 units eligible to expire from housing programs during 2025–2030.",
        "Housing units",
        "communitydistrict-eligibletoexpirefromhousingprogramsbetween2025and2030units.csv",
        "2024",
    ),
    "subsidized_units_expiring_2031_2040_cd_context": (
        "Published MN03 units eligible to expire from housing programs during 2031–2040.",
        "Housing units",
        "communitydistrict-eligibletoexpirefromhousingprogramsbetween2031and2040units.csv",
        "2024",
    ),
    "subsidized_units_expiring_2041_later_cd_context": (
        "Published MN03 units eligible to expire from housing programs in 2041 or later.",
        "Housing units",
        "communitydistrict-eligibletoexpirefromhousingprogramsin2041andlaterunits.csv",
        "2024",
    ),
}
for variable_name, values in context_dictionary.items():
    add_dictionary_row(
        variable_name,
        "Community-district context",
        values[0],
        values[1],
        "MN03 community district",
        values[2],
        values[3],
        "Select the MN 03 row and repeat the district total on each tract row for reference.",
        "This is not a tract-level value and must not be summed across rows.",
    )


# Geography, data-availability, and methodology flags.
flag_dictionary = {
    "tract_universe_flag": "Documents the official crosswalk/filter used to define the 31-tract universe.",
    "point_assignment_method": "Documents how point and property records were assigned to 2020 tracts.",
    "hpd_violations_geography_flag": "Reports the number of valid HPD records that could not be assigned to a tract.",
    "subsidized_housing_geography_flag": "Documents Furman property geography assignment and unallocated records.",
    "executed_evictions_geography_flag": "Documents eviction geography assignment and unallocated records.",
    "indoor_complaints_geography_flag": "Documents complaint geography assignment and unallocated records.",
    "nycha_violations_geography_flag": "Documents NYCHA violation geography assignment and unallocated records.",
    "eviction_filings_geography_flag": "Warns that private eviction filings are available only at community-district level.",
    "supportive_housing_status": "Documents the absence of the primary DSS/OSH supportive-housing file.",
    "new_construction_ami_status": "Documents the absence of a defensible AMI-band field in the local construction data.",
    "hpd_ll44_status": "Documents the absence of HPD Local Law 44 project and unit-income files.",
    "senior_walkup_status": "Documents the absence of MapPLUTO elevator and BIS inspection inputs.",
    "chp_2022_geography_flag": "Warns that CHP 2022 survey results are not tract-level.",
    "tenant_legal_services_geography_flag": "Documents that the provider PDF was not converted into a tract metric.",
}
for variable_name, description in flag_dictionary.items():
    add_dictionary_row(
        variable_name,
        "Methodology/status flag",
        description,
        "Text",
        "Applies to every output row",
        "Generated by this notebook",
        "Relevant source availability and geography checks",
        "Text generated from pipeline validation results or known data limitations.",
        "Interpret as metadata, not as a quantitative tract metric.",
    )

data_dictionary = pd.DataFrame(dictionary_rows)
assert set(data_dictionary["variable_name"]) == set(clean.columns)
assert data_dictionary["variable_name"].is_unique
data_dictionary = (
    data_dictionary.set_index("variable_name")
    .loc[list(clean.columns)]
    .reset_index(names="variable_name")
)
print(f"Documented {len(data_dictionary)} output variables.")

# Write the metrics and dictionary to an Excel workbook, then apply readable,
# consistent formatting. The CSV remains available as a machine-readable output.
with pd.ExcelWriter(EXCEL_OUTPUT_PATH, engine="openpyxl") as writer:
    clean.to_excel(writer, sheet_name="Metrics by Tract", index=False, startrow=3)
    data_dictionary.to_excel(writer, sheet_name="Data Dictionary", index=False, startrow=3)

workbook = load_workbook(EXCEL_OUTPUT_PATH)

primary_fill = PatternFill("solid", fgColor="1F4E78")
secondary_fill = PatternFill("solid", fgColor="D9EAF7")
accent_fill = PatternFill("solid", fgColor="E2F0D9")
warning_fill = PatternFill("solid", fgColor="FFF2CC")
white_font = Font(color="FFFFFF", bold=True)
title_font = Font(color="FFFFFF", bold=True, size=16)
subtitle_font = Font(color="404040", italic=True, size=10)
thin_gray = Side(style="thin", color="D9E1F2")
body_border = Border(bottom=thin_gray)

sheet = workbook["Metrics by Tract"]
sheet.sheet_view.showGridLines = False
sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(clean.columns))
sheet.cell(1, 1, "CB3 Housing & Affordability — Metrics by 2020 Census Tract")
sheet.cell(1, 1).fill = primary_fill
sheet.cell(1, 1).font = title_font
sheet.cell(1, 1).alignment = Alignment(horizontal="left")
sheet.row_dimensions[1].height = 26
sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(clean.columns))
sheet.cell(
    2,
    1,
    "One row per official MN03 tract. Community-district context columns repeat the same district total and must not be summed.",
)
sheet.cell(2, 1).font = subtitle_font
sheet.cell(2, 1).alignment = Alignment(wrap_text=True)
sheet.row_dimensions[2].height = 30

header_row = 4
for cell in sheet[header_row]:
    cell.fill = primary_fill
    cell.font = white_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
sheet.row_dimensions[header_row].height = 56
sheet.freeze_panes = "H5"
sheet.auto_filter.ref = f"A4:{sheet.cell(header_row, len(clean.columns)).column_letter}{4 + len(clean)}"

metrics_table = Table(
    displayName="HousingAffordabilityMetrics",
    ref=f"A4:{sheet.cell(header_row, len(clean.columns)).column_letter}{4 + len(clean)}",
)
metrics_table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
sheet.add_table(metrics_table)

# Preserve identifier fields as text and format quantitative columns by unit.
dictionary_by_variable = data_dictionary.set_index("variable_name")
for column_index, variable_name in enumerate(clean.columns, start=1):
    unit = dictionary_by_variable.loc[variable_name, "unit"]
    for row_index in range(5, 5 + len(clean)):
        cell = sheet.cell(row_index, column_index)
        cell.border = body_border
        if variable_name == "GEOID":
            cell.number_format = "@"
        elif unit == "Percent":
            cell.number_format = "0.0"
        elif unit in {
            "Households",
            "Housing units",
            "Violations",
            "Executed evictions",
            "Properties",
            "Complaints",
            "Filings",
        }:
            cell.number_format = "#,##0"

# Set bounded widths: identifiers are visible, metrics remain compact, and
# lengthy methodology/status text is readable without extreme sheet width.
for column_index, variable_name in enumerate(clean.columns, start=1):
    column_letter = sheet.cell(4, column_index).column_letter
    if variable_name == "GEOID":
        width = 14
    elif variable_name in {"tract_label", "nta_code", "cdta_code"}:
        width = 12
    elif variable_name in {"tract_name", "nta_name", "cdta_name"}:
        width = 32
    elif variable_name.endswith(("_flag", "_status")) or variable_name == "point_assignment_method":
        width = 42
        for row_index in range(5, 5 + len(clean)):
            sheet.cell(row_index, column_index).alignment = Alignment(
                vertical="top", wrap_text=True
            )
    else:
        width = 18
    sheet.column_dimensions[column_letter].width = width
for row_index in range(5, 5 + len(clean)):
    sheet.row_dimensions[row_index].height = 36


dictionary_sheet = workbook["Data Dictionary"]
dictionary_sheet.sheet_view.showGridLines = False
dictionary_sheet.merge_cells(
    start_row=1,
    start_column=1,
    end_row=1,
    end_column=len(data_dictionary.columns),
)
dictionary_sheet.cell(1, 1, "Data Dictionary — Housing & Affordability")
dictionary_sheet.cell(1, 1).fill = primary_fill
dictionary_sheet.cell(1, 1).font = title_font
dictionary_sheet.cell(1, 1).alignment = Alignment(horizontal="left")
dictionary_sheet.row_dimensions[1].height = 26
dictionary_sheet.merge_cells(
    start_row=2,
    start_column=1,
    end_row=2,
    end_column=len(data_dictionary.columns),
)
dictionary_sheet.cell(
    2,
    1,
    "Definitions follow the output column order and document source fields, calculations, geography, and limitations.",
)
dictionary_sheet.cell(2, 1).font = subtitle_font
dictionary_sheet.row_dimensions[2].height = 30

for cell in dictionary_sheet[4]:
    cell.fill = primary_fill
    cell.font = white_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
dictionary_sheet.row_dimensions[4].height = 34
dictionary_sheet.freeze_panes = "A5"

dictionary_table = Table(
    displayName="HousingAffordabilityDictionary",
    ref=f"A4:I{4 + len(data_dictionary)}",
)
dictionary_table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
dictionary_sheet.add_table(dictionary_table)

dictionary_widths = {
    "A": 45,
    "B": 28,
    "C": 55,
    "D": 20,
    "E": 24,
    "F": 48,
    "G": 50,
    "H": 70,
    "I": 65,
}
for column_letter, width in dictionary_widths.items():
    dictionary_sheet.column_dimensions[column_letter].width = width

for row in dictionary_sheet.iter_rows(
    min_row=5,
    max_row=4 + len(data_dictionary),
    min_col=1,
    max_col=len(data_dictionary.columns),
):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = body_border
    row[0].font = Font(bold=True, color="1F1F1F")
    if row[1].value == "Community-district context":
        for cell in row:
            cell.fill = warning_fill
    elif row[1].value == "Methodology/status flag":
        for cell in row:
            cell.fill = secondary_fill
    elif row[1].value == "Geography":
        for cell in row:
            cell.fill = accent_fill
    dictionary_sheet.row_dimensions[row[0].row].height = 72

workbook.save(EXCEL_OUTPUT_PATH)
print(f"Wrote Excel workbook to {EXCEL_OUTPUT_PATH}")
